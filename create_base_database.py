#!/usr/bin/env python3
"""Create base Excel database file for the budget tracker"""

import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_base_database():
    """Create a properly formatted Excel file to serve as database"""
    
    # Create empty DataFrame with proper structure
    df = pd.DataFrame(columns=[
        'id', 'type', 'amount', 'description', 'category', 'date'
    ])
    
    # Save to Excel
    filename = 'budget_data.xlsx'
    df.to_excel(filename, index=False, sheet_name='Transactions')
    
    # Open with openpyxl to format properly
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook['Transactions']
    
    # Define styles
    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    headers = ['ID', 'Type', 'Amount (₱)', 'Description', 'Category', 'Date']
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths for better text fitting
    column_widths = {
        'A': 8,   # ID
        'B': 12,  # Type
        'C': 15,  # Amount
        'D': 30,  # Description
        'E': 15,  # Category
        'F': 12   # Date
    }
    
    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width
    
    # Set row height
    worksheet.row_dimensions[1].height = 20
    
    # Apply auto-filter
    worksheet.auto_filter.ref = f"A1:F1"
    
    # Freeze the header row
    worksheet.freeze_panes = "A2"
    
    # Save the formatted workbook
    workbook.save(filename)
    
    print(f"Base database file '{filename}' created successfully!")
    print("Structure:")
    print("   - ID: Auto-incremented unique identifier")
    print("   - Type: income or expense")  
    print("   - Amount: Monetary value in Philippine Peso")
    print("   - Description: Transaction details")
    print("   - Category: Transaction category")
    print("   - Date: Transaction date")
    print("\nFeatures:")
    print("   - Formatted headers with proper styling")
    print("   - Optimized column widths")
    print("   - Auto-filter enabled")
    print("   - Header row frozen for easy navigation")

if __name__ == "__main__":
    create_base_database()