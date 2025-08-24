import os
import pandas as pd
from datetime import datetime, date
from typing import List, Dict, Optional
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openai import OpenAI
from dotenv import load_dotenv
import streamlit as st

load_dotenv()

class BudgetTrackerWeb:
    def __init__(self, excel_file: str = 'budget_data.xlsx'):
        self.excel_file = excel_file
        self.user_profile_file = 'user_profile.xlsx'
        self.client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        self._ensure_database_integrity()
        
    def _ensure_database_integrity(self):
        """Ensure database files exist and have proper structure"""
        # Create transactions file if it doesn't exist
        if not os.path.exists(self.excel_file):
            self._create_empty_transactions_file()
        
        # Create user profile file if it doesn't exist
        if not os.path.exists(self.user_profile_file):
            self._create_empty_user_profile()
        
        # Validate and repair data integrity
        self._validate_data_integrity()
    
    def _create_empty_transactions_file(self):
        """Create empty transactions Excel file with proper structure"""
        df = pd.DataFrame(columns=['id', 'type', 'amount', 'description', 'category', 'date'])
        self.save_data(df)
    
    def _create_empty_user_profile(self):
        """Create empty user profile Excel file"""
        profile_data = {
            'setting': ['monthly_income', 'savings_goal', 'expense_limit', 'setup_completed', 'created_date'],
            'value': [0, 0, 0, False, datetime.now().strftime('%Y-%m-%d')]
        }
        df = pd.DataFrame(profile_data)
        df.to_excel(self.user_profile_file, index=False, sheet_name='UserProfile')
    
    def _validate_data_integrity(self):
        """Validate and ensure data integrity"""
        try:
            df = self.load_data()
            if not df.empty:
                # Ensure ID uniqueness and proper sequencing
                if df['id'].duplicated().any():
                    # Fix duplicate IDs
                    df['id'] = range(1, len(df) + 1)
                    self.save_data(df)
                
                # Ensure proper data types
                df['amount'] = pd.to_numeric(df['amount'], errors='coerce').fillna(0)
                df['date'] = pd.to_datetime(df['date'], errors='coerce').fillna(datetime.now())
                
                # Ensure valid transaction types
                df['type'] = df['type'].apply(lambda x: x.lower() if x.lower() in ['income', 'expense'] else 'expense')
                
                self.save_data(df)
        except Exception as e:
            st.warning(f"Data integrity check warning: {str(e)}")
    
    def load_data(self, date_filter: str = None) -> pd.DataFrame:
        """Load data from Excel file with optional date filtering"""
        try:
            if os.path.exists(self.excel_file):
                df = pd.read_excel(self.excel_file)
                # Ensure date column is datetime
                if 'date' in df.columns and len(df) > 0:
                    df['date'] = pd.to_datetime(df['date'])
                    
                    # Apply date filter if specified
                    if date_filter == 'current_month':
                        current_date = datetime.now()
                        df = df[(df['date'].dt.year == current_date.year) & 
                               (df['date'].dt.month == current_date.month)]
                    elif date_filter == 'current_year':
                        current_year = datetime.now().year
                        df = df[df['date'].dt.year == current_year]
                
                return df
            else:
                # Create empty DataFrame with required columns
                return pd.DataFrame(columns=['id', 'type', 'amount', 'description', 'category', 'date'])
        except Exception as e:
            st.error(f"Error loading data: {str(e)}")
            return pd.DataFrame(columns=['id', 'type', 'amount', 'description', 'category', 'date'])
    
    def save_data(self, df: pd.DataFrame):
        """Save data to Excel file with proper formatting"""
        try:
            # Save to Excel
            df.to_excel(self.excel_file, index=False)
            
            # Apply formatting if openpyxl is available
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                workbook = openpyxl.load_workbook(self.excel_file)
                worksheet = workbook.active
                
                # Format headers if data exists
                if len(df) >= 0:  # Always format headers
                    header_font = Font(bold=True, size=12, color='FFFFFF')
                    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                    header_alignment = Alignment(horizontal='center', vertical='center')
                    
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Format header row
                    for col_idx in range(1, 7):  # 6 columns
                        cell = worksheet.cell(row=1, column=col_idx)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    # Set column widths
                    column_widths = ['A:8', 'B:12', 'C:15', 'D:30', 'E:15', 'F:12']
                    for col_width in column_widths:
                        col, width = col_width.split(':')
                        worksheet.column_dimensions[col].width = int(width)
                
                workbook.save(self.excel_file)
            except ImportError:
                pass  # Skip formatting if openpyxl not available
                
        except Exception as e:
            st.error(f"Error saving data: {str(e)}")
    
    def add_transaction(self, transaction_type: str, amount: float, description: str, 
                       category: str = None, date_input: date = None) -> pd.DataFrame:
        """Add a new transaction with proper ID management"""
        df = self.load_data()
        
        # Generate new ID - find the maximum existing ID and add 1
        if len(df) > 0 and 'id' in df.columns:
            new_id = int(df['id'].max()) + 1
        else:
            new_id = 1
        
        # Use AI categorization for expenses if no category provided
        if transaction_type == 'expense' and not category:
            category = self.ai_categorize_expense(description)
        elif transaction_type == 'income':
            category = 'Income'
        
        new_transaction = {
            'id': new_id,
            'type': transaction_type,
            'amount': amount,
            'description': description,
            'category': category,
            'date': date_input or datetime.now().date()
        }
        
        # Append transaction to existing data
        new_df = pd.concat([df, pd.DataFrame([new_transaction])], ignore_index=True)
        
        # Sort by ID to maintain order
        new_df = new_df.sort_values('id').reset_index(drop=True)
        
        self.save_data(new_df)
        return new_df
    
    def delete_transaction(self, transaction_id: int) -> pd.DataFrame:
        """Delete a transaction by ID"""
        df = self.load_data()
        
        if df.empty:
            raise ValueError("No transactions to delete")
        
        if transaction_id not in df['id'].values:
            raise ValueError(f"Transaction with ID {transaction_id} not found")
        
        # Remove the transaction
        new_df = df[df['id'] != transaction_id].copy()
        
        # Reset index but keep original IDs
        new_df = new_df.reset_index(drop=True)
        
        self.save_data(new_df)
        return new_df
    
    def get_transaction_by_id(self, transaction_id: int) -> pd.Series:
        """Get a specific transaction by ID"""
        df = self.load_data()
        
        if df.empty or transaction_id not in df['id'].values:
            return None
        
        return df[df['id'] == transaction_id].iloc[0]
    
    def ai_categorize_expense(self, description: str) -> str:
        """Use AI to categorize expense"""
        try:
            if not os.getenv('OPENAI_API_KEY'):
                return "Other"
                
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a financial categorization assistant. Categorize the expense into one of these categories: Food, Transportation, Entertainment, Healthcare, Shopping, Utilities, Housing, Education, Other. Return only the category name."},
                    {"role": "user", "content": f"Categorize this expense: {description}"}
                ]
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            st.warning("AI categorization unavailable, using 'Other'")
            return "Other"
    
    def get_balance(self, df: pd.DataFrame) -> float:
        """Calculate current balance"""
        if df.empty:
            return 0.0
        income = df[df['type'] == 'income']['amount'].sum()
        expenses = df[df['type'] == 'expense']['amount'].sum()
        return income - expenses
    
    def get_monthly_summary(self, df: pd.DataFrame, year: int = None, month: int = None) -> Dict:
        """Get monthly financial summary"""
        if df.empty:
            return {'income': 0, 'expenses': 0, 'balance': 0, 'expense_by_category': {}, 'transaction_count': 0}
        
        if not year:
            year = datetime.now().year
        if not month:
            month = datetime.now().month
            
        # Filter for the specific month
        monthly_df = df[
            (df['date'].dt.year == year) & 
            (df['date'].dt.month == month)
        ].copy()
        
        if monthly_df.empty:
            return {'income': 0, 'expenses': 0, 'balance': 0, 'expense_by_category': {}, 'transaction_count': 0}
        
        income = monthly_df[monthly_df['type'] == 'income']['amount'].sum()
        expenses = monthly_df[monthly_df['type'] == 'expense']['amount'].sum()
        
        # Group expenses by category
        expense_by_category = monthly_df[monthly_df['type'] == 'expense'].groupby('category')['amount'].sum().to_dict()
        
        return {
            'income': income,
            'expenses': expenses,
            'balance': income - expenses,
            'expense_by_category': expense_by_category,
            'transaction_count': len(monthly_df)
        }
    
    def create_expense_pie_chart(self, df: pd.DataFrame):
        """Create pie chart of expenses by category"""
        if df.empty:
            return None
        
        expenses = df[df['type'] == 'expense']
        if expenses.empty:
            return None
        
        category_totals = expenses.groupby('category')['amount'].sum().reset_index()
        
        fig = px.pie(category_totals, values='amount', names='category', 
                    title='Expenses by Category')
        return fig
    
    def create_monthly_trend_chart(self, df: pd.DataFrame):
        """Create monthly trend chart"""
        if df.empty:
            return None
        
        # Group by month and type
        df['year_month'] = df['date'].dt.to_period('M').astype(str)
        monthly_data = df.groupby(['year_month', 'type'])['amount'].sum().reset_index()
        
        fig = px.line(monthly_data, x='year_month', y='amount', color='type',
                     title='Monthly Income vs Expenses Trend',
                     labels={'year_month': 'Month', 'amount': 'Amount (₱)'})
        return fig
    
    def create_balance_chart(self, df: pd.DataFrame):
        """Create running balance chart"""
        if df.empty:
            return None
        
        # Sort by date
        df_sorted = df.sort_values('date').copy()
        
        # Calculate running balance
        df_sorted['running_balance'] = 0.0
        balance = 0.0
        
        for idx, row in df_sorted.iterrows():
            if row['type'] == 'income':
                balance += row['amount']
            else:
                balance -= row['amount']
            df_sorted.at[idx, 'running_balance'] = balance
        
        fig = px.line(df_sorted, x='date', y='running_balance',
                     title='Running Balance Over Time',
                     labels={'date': 'Date', 'running_balance': 'Balance (₱)'})
        return fig
    
    def create_category_bar_chart(self, df: pd.DataFrame):
        """Create horizontal bar chart of spending by category"""
        if df.empty:
            return None
        
        expenses = df[df['type'] == 'expense']
        if expenses.empty:
            return None
        
        category_totals = expenses.groupby('category')['amount'].sum().sort_values(ascending=True)
        
        fig = px.bar(x=category_totals.values, y=category_totals.index,
                    orientation='h',
                    title='Spending by Category',
                    labels={'x': 'Amount (₱)', 'y': 'Category'})
        return fig
    
    def ai_spending_analysis(self, df: pd.DataFrame) -> str:
        """Get AI-powered spending analysis"""
        if df.empty or not os.getenv('OPENAI_API_KEY'):
            return "No data available for analysis or OpenAI API key not configured."
        
        summary = self.get_monthly_summary(df)
        recent_expenses = df[df['type'] == 'expense'].tail(10).to_dict('records')
        
        analysis_data = {
            'monthly_summary': summary,
            'recent_expenses': recent_expenses,
            'balance': self.get_balance(df)
        }
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a financial advisor. Analyze the spending data and provide insights, patterns, and recommendations. Be concise but helpful."},
                    {"role": "user", "content": f"Analyze this financial data: {str(analysis_data)}"}
                ]
            )
            return response.choices[0].message.content
        except Exception as e:
            return f"AI analysis unavailable: {str(e)}"
    
    def ai_budget_recommendations(self, df: pd.DataFrame, monthly_income: float) -> str:
        """Get AI-powered budget recommendations"""
        if df.empty or not os.getenv('OPENAI_API_KEY'):
            return "No data available for recommendations or OpenAI API key not configured."
        
        summary = self.get_monthly_summary(df)
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a financial advisor. Based on income and spending patterns, provide budget recommendations using the 50/30/20 rule or other appropriate strategies."},
                    {"role": "user", "content": f"Monthly income: ₱{monthly_income}, Current spending: {str(summary)}"}
                ]
            )
            return response.choices[0].message.content
        except Exception as e:
            return f"Budget recommendations unavailable: {str(e)}"
    
    def load_user_profile(self) -> Dict:
        """Load user profile settings"""
        try:
            if os.path.exists(self.user_profile_file):
                df = pd.read_excel(self.user_profile_file)
                profile = {}
                for _, row in df.iterrows():
                    profile[row['setting']] = row['value']
                return profile
            else:
                return {
                    'monthly_income': 0,
                    'savings_goal': 0,
                    'expense_limit': 0,
                    'setup_completed': False,
                    'created_date': datetime.now().strftime('%Y-%m-%d')
                }
        except Exception as e:
            st.error(f"Error loading user profile: {str(e)}")
            return {'monthly_income': 0, 'savings_goal': 0, 'expense_limit': 0, 'setup_completed': False}
    
    def save_user_profile(self, profile_data: Dict):
        """Save user profile settings"""
        try:
            data = []
            for key, value in profile_data.items():
                data.append({'setting': key, 'value': value})
            
            df = pd.DataFrame(data)
            df.to_excel(self.user_profile_file, index=False, sheet_name='UserProfile')
        except Exception as e:
            st.error(f"Error saving user profile: {str(e)}")
    
    def is_first_time_user(self) -> bool:
        """Check if this is a first-time user"""
        profile = self.load_user_profile()
        return not profile.get('setup_completed', False)
    
    def get_financial_overview(self, df: pd.DataFrame, view_type: str = 'all') -> Dict:
        """Get comprehensive financial overview with different view types"""
        if df.empty:
            return {
                'total_balance': 0,
                'total_income': 0,
                'total_expenses': 0,
                'avg_monthly_income': 0,
                'avg_monthly_expenses': 0,
                'largest_expense': 0,
                'most_frequent_category': 'N/A',
                'transaction_count': 0,
                'date_range': 'No data'
            }
        
        # Filter data based on view type
        if view_type == 'current_month':
            current_date = datetime.now()
            filtered_df = df[(df['date'].dt.year == current_date.year) & 
                           (df['date'].dt.month == current_date.month)]
        elif view_type == 'current_year':
            current_year = datetime.now().year
            filtered_df = df[df['date'].dt.year == current_year]
        else:
            filtered_df = df.copy()
        
        if filtered_df.empty:
            return self.get_financial_overview(pd.DataFrame(columns=df.columns), 'all')
        
        income_df = filtered_df[filtered_df['type'] == 'income']
        expense_df = filtered_df[filtered_df['type'] == 'expense']
        
        total_income = income_df['amount'].sum() if not income_df.empty else 0
        total_expenses = expense_df['amount'].sum() if not expense_df.empty else 0
        
        # Calculate monthly averages
        date_range = (filtered_df['date'].max() - filtered_df['date'].min()).days
        months_span = max(1, date_range / 30.44)  # Average days per month
        
        return {
            'total_balance': total_income - total_expenses,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'avg_monthly_income': total_income / months_span,
            'avg_monthly_expenses': total_expenses / months_span,
            'largest_expense': expense_df['amount'].max() if not expense_df.empty else 0,
            'most_frequent_category': expense_df['category'].mode().iloc[0] if not expense_df.empty and not expense_df['category'].mode().empty else 'N/A',
            'transaction_count': len(filtered_df),
            'date_range': f"{filtered_df['date'].min().strftime('%Y-%m-%d')} to {filtered_df['date'].max().strftime('%Y-%m-%d')}"
        }